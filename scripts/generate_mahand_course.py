#!/usr/bin/env python3
"""Generate the Mahand course TypeScript snapshot from the source Excel workbook.

Usage:
  python scripts/generate_mahand_course.py /path/to/mahand.xlsx

The workbook must contain a flat vocabulary table with columns for unit/group/page,
word number, English term, Arabic translation, English example, and Arabic example.
Header aliases in Arabic and English are accepted.
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path
import re

from openpyxl import load_workbook

EXPECTED_UNITS = 20
EXPECTED_GROUPS = 100
EXPECTED_ROWS = 1041

ALIASES = {
    "unit_no": {"unit", "unit no", "unit number", "unit_no", "رقم الوحدة", "الوحدة رقم"},
    "unit_title": {"unit title", "unit name", "unit_title", "اسم الوحدة", "عنوان الوحدة", "الوحدة"},
    "group_no": {"group", "group no", "group number", "group_no", "رقم الجروب", "رقم المجموعة"},
    "group_title": {"group title", "group name", "group_title", "اسم الجروب", "اسم المجموعة", "عنوان الجروب"},
    "page": {"page", "page no", "page number", "pdf page", "صفحة", "رقم الصفحة"},
    "word_no": {"word no", "word number", "item no", "number", "word_no", "رقم الكلمة", "م"},
    "term": {"word", "term", "english", "english word", "vocabulary", "الكلمة", "انجليزي", "الانجليزية"},
    "translation": {"translation", "meaning", "arabic", "arabic meaning", "الترجمة", "المعنى", "عربي"},
    "example": {"example", "sentence", "example sentence", "english example", "مثال", "الجملة", "مثال انجليزي"},
    "example_translation": {"example translation", "sentence translation", "arabic example", "arabic sentence", "ترجمة المثال", "ترجمة الجملة", "مثال عربي"},
}

REQUIRED = tuple(ALIASES)


def norm(value: object) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"\s+", " ", text.strip().lower())
    return text.replace("_", " ")


def find_table(workbook):
    for sheet in workbook.worksheets:
        for row_no, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(25, sheet.max_row), values_only=True), start=1):
            normalized = [norm(cell) for cell in row]
            mapping: dict[str, int] = {}
            for key, aliases in ALIASES.items():
                normalized_aliases = {norm(alias) for alias in aliases}
                for index, heading in enumerate(normalized):
                    if heading in normalized_aliases:
                        mapping[key] = index
                        break
            if all(key in mapping for key in REQUIRED):
                return sheet, row_no, mapping
    raise SystemExit("Could not find a vocabulary sheet with the required headers.")


def as_int(value: object, field: str, row_no: int) -> int:
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        raise SystemExit(f"Invalid {field} at Excel row {row_no}: {value!r}") from None


def clean(value: object) -> str:
    return "" if value is None else re.sub(r"\s+", " ", str(value).strip())


def ts_escape(value: str) -> str:
    return value.replace("\\", "\\\\").replace("`", "\\`").replace("${", "\\${").replace("\t", " ").replace("\n", " ")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output-dir", type=Path, default=Path("src/curriculum/mahand/source"))
    parser.add_argument("--allow-count-drift", action="store_true")
    args = parser.parse_args()

    wb = load_workbook(args.workbook, read_only=True, data_only=True)
    sheet, header_row, columns = find_table(wb)

    rows: list[tuple[int, str, int, str, int, int, str, str, str, str]] = []
    groups: set[tuple[int, int]] = set()
    units: set[int] = set()

    for excel_row_no, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        term = clean(values[columns["term"]] if columns["term"] < len(values) else None)
        if not term:
            continue
        unit_no = as_int(values[columns["unit_no"]], "unit number", excel_row_no)
        group_no = as_int(values[columns["group_no"]], "group number", excel_row_no)
        page = as_int(values[columns["page"]], "page", excel_row_no)
        word_no = as_int(values[columns["word_no"]], "word number", excel_row_no)
        record = (
            unit_no,
            clean(values[columns["unit_title"]]),
            group_no,
            clean(values[columns["group_title"]]),
            page,
            word_no,
            term,
            clean(values[columns["translation"]]),
            clean(values[columns["example"]]),
            clean(values[columns["example_translation"]]),
        )
        rows.append(record)
        units.add(unit_no)
        groups.add((unit_no, group_no))

    counts = (len(units), len(groups), len(rows))
    expected = (EXPECTED_UNITS, EXPECTED_GROUPS, EXPECTED_ROWS)
    if counts != expected and not args.allow_count_drift:
        raise SystemExit(f"Unexpected Mahand counts: got {counts}, expected {expected}. Use --allow-count-drift to override.")

    by_pair: dict[tuple[int, int], list[tuple[int, str, int, str, int, int, str, str, str, str]]] = defaultdict(list)
    for row in rows:
        pair_start = ((row[0] - 1) // 2) * 2 + 1
        by_pair[(pair_start, pair_start + 1)].append(row)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    for (start, end), pair_rows in sorted(by_pair.items()):
        const_name = f"MAHAND_ROWS_{start:02d}_{end:02d}"
        lines = ["\t".join(ts_escape(str(value)) for value in row) for row in pair_rows]
        content = (
            "// Generated by scripts/generate_mahand_course.py. Do not edit by hand.\n"
            f"export const {const_name} = `" + "\n".join(lines) + "`;\n"
        )
        (args.output_dir / f"units-{start:02d}-{end:02d}.ts").write_text(content, encoding="utf-8")

    print(f"Generated Mahand: {len(units)} units, {len(groups)} groups, {len(rows)} vocabulary rows from {sheet.title!r}.")


if __name__ == "__main__":
    main()
